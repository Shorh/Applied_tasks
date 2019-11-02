import pandas as pd
import openpyxl

from openpyxl.styles import Alignment, Font, PatternFill
from data import constants
from data.table_create import ReportTables


class WriteTable:
    def __init__(self, file_name, first_sheet_name):
        self.file_name = file_name
        self.first_sheet_name = first_sheet_name
        self.work_book = None
        self.sheet = None
        self.__create_book()

        self.sheet_names = self.__get_sheets()
        self.row = 1

    def __create_book(self):
        self.work_book = openpyxl.Workbook()
        self.sheet = self.work_book.active
        self.sheet.title = self.first_sheet_name
        self.work_book.save(self.file_name)

    def __get_sheets(self):
        return dict((ws.title, ws) for ws in self.work_book.worksheets)

    def get_writer(self):
        writer = pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a',
                                datetime_format='DD.MM.YYYY')
        self.sheet_names = self.__get_sheets()
        return writer

    def write_titles(self, type_table, **kwargs):
        """
        Запись в Excel описания таблицы
        :param type_table: тип описания
        1. Ведомость по ЗП. Первая страница
        2. Табель по объекту
        3. Валовая прибыль (за прошлый период или за месяц)
        :param sheet: ссылка на страницу
        :param kwargs: Если type_table == 1: city
        Если type_table == 2: month, year, partner, obj_name, address
        Если type_table == 3: None или month
        :return: None
        """
        if type_table == 1:
            city = kwargs['city']
            self.row = 1
            self.sheet.cell(row=self.row, column=1).value = f'ВЕДОМОСТЬ ПО ЗП'
            self.sheet.cell(row=self.row, column=1).font = constants.FONT_TITLE
            self.row += 1

            self.sheet.cell(row=self.row, column=1).value = f'Филиал'
            self.sheet.cell(row=self.row, column=2).value = city.upper()
            self.sheet.cell(row=self.row, column=2).font = constants.FONT_TITLE
            self.row += 1

            self.sheet.cell(row=self.row, column=1).value = f'Дата формирования отчета: '
            self.sheet.cell(row=self.row, column=2).value = f'{constants.TODAY:%d.%m.%Y}'
            self.row += 1

            self.sheet.cell(row=self.row, column=1).value = f'Отчетные даты: '
            self.sheet.cell(row=self.row, column=2).value = \
                f'{constants.START_LAST_WEEK:%d.%m.%Y} - ' \
                f'{constants.END_LAST_WEEK:%d.%m.%Y}'
            self.row += 1

        elif type_table == 2:
            month = kwargs['month']
            year = kwargs['year']
            partner = kwargs['partner']
            obj_name = kwargs['obj_name']
            address = kwargs['address']

            self.row = 1
            self.sheet.cell(row=self.row, column=1).value = \
                f'Лист учета рабочего времени {constants.MONTHS[month]} {year}'
            self.sheet.cell(row=self.row, column=1).font = constants.FONT_TITLE

            self.row += 1
            self.sheet.cell(row=self.row, column=1).value = f'Дата формирования отчета: '
            self.sheet.cell(row=self.row, column=2).value = f'{constants.TODAY:%d.%m.%Y}'

            self.row += 1
            self.sheet.cell(row=self.row, column=1).value = 'ПАРТНЕР'
            self.sheet.cell(row=self.row, column=1).font = constants.FONT_TITLE
            self.sheet.cell(row=self.row, column=2).value = partner

            self.row += 1
            self.sheet.cell(row=self.row, column=1).value = 'ОБЪЕКТ'
            self.sheet.cell(row=self.row, column=1).font = constants.FONT_TITLE
            self.sheet.cell(row=self.row, column=2).value = obj_name

            self.row += 1
            self.sheet.cell(row=self.row, column=1).value = 'АДРЕС ОБЪЕКТА'
            self.sheet.cell(row=self.row, column=1).font = constants.FONT_TITLE
            self.sheet.cell(row=self.row, column=2).value = address
            self.row += 1

        elif type_table == 3:
            self.row = 1
            self.sheet.cell(row=self.row, column=1).value = f'ВАЛОВАЯ ПРИБЫЛЬ'
            if 'month' in kwargs:
                self.sheet.cell(row=self.row, column=1).value += \
                    f'. {constants.MONTHS[kwargs["month"]]}'
            self.sheet.cell(row=self.row, column=1).font = constants.FONT_TITLE

            self.row += 1
            self.sheet.cell(row=self.row, column=1).value = f'Дата формирования отчета: '
            self.sheet.cell(row=self.row, column=2).value = f'{constants.TODAY:%d.%m.%Y}'

            self.row += 1
            if 'month' not in kwargs:
                self.sheet.cell(row=self.row, column=1).value = f'Отчетные даты: '
                self.sheet.cell(row=self.row, column=2).value = \
                    f'{constants.START_LAST_WEEK:%d.%m.%Y} - ' \
                    f'{constants.END_LAST_WEEK:%d.%m.%Y}'
                self.row += 1

            for i, _ in enumerate(self.sheet):
                self.sheet.cell(row=i + 1, column=1).alignment = Alignment(
                    horizontal='left')

    @staticmethod
    def add_result(sheet, row, *args):
        """
        Добавление итоговой строки таблицы
        :param sheet: ссылка на страницу Excel
        :param row: номер строки, куда необходимо вставить результат
        :param args: параметры итоговой строки, type - кортеж кортежей
        :return: None
        """
        for column, value in args:
            sheet.cell(row=row, column=column).value = value

    def write_table(self, writer, table, **kwargs):
        """
        Печать в файл Excel таблиц зарплатной ведомости
        :param writer: контектсный менеджер pandas записи в Excel
        :param table: ссылка на таблицу
        :param kwargs: table_res: параметры итоговой строки, type - кортеж кортежей.
        table_name: имя таблицы. По умолчанию - None.
        is_auto_dimension: необходима ли автоширина столбцов таблицы в файле Excel
        :return: None
        """
        if table.shape[0] != 0:
            if 'table_name' in kwargs:
                self.row += 2
                self.sheet.cell(row=self.row, column=1).value = kwargs['table_name']
                self.sheet.cell(row=self.row, column=1).font = constants.FONT_TITLE
            self.row += 1

            table.to_excel(writer, self.sheet.title,
                           index=False, startrow=self.row - 1)
            if 'is_auto_dimension' in kwargs and kwargs['is_auto_dimension']:
                self.style_auto_dimension(column_count=3)
                self.style_dimension(
                    column_numbers=list(range(4, table.shape[1] + 1)))

            if 'table_res' in kwargs:
                self.row += table.shape[0] + 1
                self.add_result(self.sheet, self.row, *kwargs['table_res'])

            self.style_table(row_start=self.row - table.shape[0] - 1,
                             table_size=(table.shape[0] + 2, table.shape[1]))

    def style_auto_dimension(self, **kwargs):
        """
        Автоширина столбцов таблицы в файле Excel
        :param kwargs: column_count: Количество колонок для автоширины.
        По умолчанию - все колонки на странице.
        add_width: добавить к ширине столбца символов. По умолчанию - 4.
        is_formatting: форматировать значения ячеек как числа? По умолчанию - True
        :return: None
        """
        if 'column_count' not in kwargs:
            kwargs['column_count'] = len(self.sheet[self.row - 1])
        if 'add_width' not in kwargs:
            kwargs['add_width'] = 4
        if 'is_formatting' not in kwargs:
            kwargs['is_formatting'] = True

        column_widths = []
        for i, row in enumerate(self.sheet):
            if i < self.row - 1:
                continue
            for j, cell in enumerate(row):
                if j >= kwargs['column_count']:
                    break
                if kwargs['is_formatting']:
                    cell.number_format = constants.NUMBER_FORMAT
                if len(column_widths) > j:
                    if cell.value and len(str(cell.value)) > column_widths[j]:
                        column_widths[j] = len(str(cell.value))
                else:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            self.sheet.column_dimensions[
                openpyxl.utils.get_column_letter(
                    i + 1)].width = column_width + kwargs['add_width']

    def style_dimension(self, width=12, column_numbers=None):
        """
        Ширина столбцов таблицы в файле Excel
        :param width: ширина столбца
        :param column_numbers: номера столбцов для изменения ширины
        :return: None
        """
        if column_numbers:
            for i in column_numbers:
                self.sheet.column_dimensions[
                    openpyxl.utils.get_column_letter(i)].width = width

    def style_result(self, column_count, **kwargs):
        """
        Форматирование стартовых и итоговых строк таблицы
        :param column_count: количество столбцов
        :param kwargs: row: ссылка на строку Excel. По умолчанию - текущая строка.
        last_res: является ли строка итоговой? По умолчанию - False.
        alignment: выравнивание, по умолчанию - центр.
        is_formatting: форматировать значения ячеек как числа? По умолчанию - True
        :return: None
        """
        if 'row' not in kwargs:
            kwargs['row'] = self.sheet[self.row]
        if 'is_formatting' not in kwargs:
            kwargs['is_formatting'] = True
        if 'last_res' not in kwargs:
            kwargs['last_res'] = False
        if 'alignment' not in kwargs:
            kwargs['alignment'] = 'center'

        for i, cell in enumerate(kwargs['row']):
            if i >= column_count:
                break
            if kwargs['is_formatting']:
                cell.number_format = constants.NUMBER_FORMAT

            cell.fill = constants.RESULT_FILL
            cell.font = Font(bold=True)
            cell.border = constants.BORDER

            if not kwargs['last_res']:
                cell.alignment = Alignment(horizontal=kwargs['alignment'],
                                           vertical='center', wrapText=True)
            else:
                cell.alignment = Alignment(horizontal=kwargs['alignment'],
                                           vertical='center')
        if kwargs['last_res']:
            kwargs['row'][0].alignment = Alignment(horizontal='right')

    def style_table(self, row_start, **kwargs):
        """
        Формат таблицы: границы, формат числа и даты, стартовая и итоговая строки
        :param row_start: номер стартовой строки
        :param kwargs: table_size: (кол-во строк, кол-во столбцов). По умолчанию (1, 1).
        is_formatting: форматировать значения ячеек как числа и даты? По умолчанию - True.
        is_result_style: форматировать ли стартовую и итоговую строки таблицы? По умолчанию (True, True)
        :return: None
        """
        if 'table_size' not in kwargs:
            kwargs['table_size'] = (1, 1)
        if 'is_formatting' not in kwargs:
            kwargs['is_formatting'] = True
        if 'is_result_style' not in kwargs:
            kwargs['is_result_style'] = (True, True)

        row_end = row_start + kwargs['table_size'][0] - 1
        column_count = kwargs['table_size'][1]

        for row in range(row_start, row_end + 1):
            for i, cell in enumerate(self.sheet[row]):
                if i >= column_count:
                    break
                if kwargs['is_formatting']:
                    if 'Дата' in self.sheet.cell(row=row_start, column=i + 1).value:
                        cell.number_format = constants.DATA_FORMAT
                    else:
                        cell.number_format = constants.NUMBER_FORMAT
                cell.border = constants.BORDER
        if kwargs['is_result_style'][0]:
            self.style_result(column_count=column_count,
                              row=self.sheet[row_start],
                              is_formatting=kwargs['is_formatting'])
        if kwargs['is_result_style'][1]:
            self.style_result(column_count=column_count,
                              row=self.sheet[row_end], last_res=True,
                              is_formatting=kwargs['is_formatting'])


class WriteReportTables(WriteTable):
    def write_weekly_report(self, city, report_tables):
        salary_record = report_tables['record']
        salary_fine_detail = report_tables['fine']
        salary_done_detail = report_tables['done']

        self.work_book = openpyxl.load_workbook(self.file_name)
        self.sheet = self.work_book[self.first_sheet_name]
        self.row = 1

        with self.get_writer() as writer:
            writer.book = self.work_book
            writer.sheets = self.sheet_names
            self.write_titles(type_table=1, city=city)

            res = ((1, 'ИТОГО'),
                   (4, sum(salary_record['Начислено'])),
                   (5, sum(salary_record['Выплачено'])),
                   (6, sum(salary_record['Долг'])),
                   (7, sum(salary_record['Штраф/премия текущей недели'])),
                   (8, sum(salary_record['К выплате'])))
            self.write_table(writer, salary_record.iloc[:, 1:], table_res=res,
                             is_auto_dimension=True)

            if salary_fine_detail.shape[0] != 0:
                res = ((1, 'ИТОГО'),
                       (5, sum(salary_fine_detail['Штраф'])))
                self.write_table(writer, salary_fine_detail.iloc[:, 1:],
                                 table_res=res,
                                 table_name='ШТРАФЫ ТЕКУЩЕЙ НЕДЕЛИ')

            if salary_done_detail.shape[0] != 0:
                res = ((1, 'ИТОГО'),
                       (4, sum(salary_done_detail['Сумма выплаты'])))
                self.write_table(writer, salary_done_detail.iloc[:, 1:],
                                 table_res=res, table_name='ВЫПЛАТЫ')

        self.work_book.save(self.file_name)

    def write_month_report(self, tables_for_report, obj_id, month):
        """
        Печать табеля в файл Excel
        :param tables_for_report: словарь таблиц для табеля, содержащий
        shift_for_report и done_for_report
        :param obj_id: id объекта
        :param month: номер месяца
        :return: None
        """
        if month == constants.MONTH_START:
            year = constants.YEAR_START
        else:
            year = constants.YEAR_END

        shift_for_report = tables_for_report['shift_for_report']
        done_for_report = tables_for_report['done_for_report']

        table_shift = shift_for_report[(shift_for_report['ObjectId'] == obj_id) &
                                       (shift_for_report['month'] == month)]
        table_done = done_for_report[(done_for_report['ObjectId'] == obj_id) &
                                     (done_for_report['month'] == month)]

        month_report_obj = \
            ReportTables().get_month_report(table_shift, month, year)

        group_by = ['Partner', 'PartnerName', 'ObjectId',
                    'ObjectName', 'ObjectAddress']
        shift_for_report_unique = table_shift[group_by].drop_duplicates().reset_index()

        obj_name = \
            shift_for_report_unique[shift_for_report_unique['ObjectId'] ==
                                    obj_id]['ObjectName'].values[0]
        partner = \
            shift_for_report_unique[shift_for_report_unique['ObjectId'] ==
                                    obj_id]['PartnerName'].values[0]
        address = \
            shift_for_report_unique[shift_for_report_unique['ObjectId'] ==
                                    obj_id]['ObjectAddress'].values[0]
        sheet_name = f'shift_report_{month}_{int(obj_id)}'

        self.work_book = openpyxl.load_workbook(self.file_name)
        self.work_book.create_sheet(sheet_name)
        self.sheet = self.work_book[sheet_name]
        self.row = 1

        with self.get_writer() as writer:
            writer.book = self.work_book
            writer.sheets = self.sheet_names
            self.write_titles(type_table=2, month=month, year=year,
                              partner=partner, obj_name=obj_name,
                              address=address)

            month_report_obj.to_excel(writer, sheet_name, startrow=self.row)
            self.row += 1
            self.sheet.cell(row=self.row, column=1).value = 'ФИО сотрудника'

            self.sheet.insert_rows(self.row + 1)
            for i, cell in enumerate(self.sheet[self.row + 1]):
                if i < 2:
                    continue
                weekday = int(f'{self.sheet.cell(row=self.row, column=i + 1).value:%w}')
                weekday = constants.WEEKDAYS[weekday]
                cell.value = weekday

            for i, _ in enumerate(self.sheet):
                self.sheet.cell(row=i + 1, column=1).alignment = \
                    Alignment(horizontal='left')

            res = [(1, 'ИТОГО ЧАСОВ')] + list(
                (i + 2, sum(month_report_obj[column])) for i, column in
                enumerate(month_report_obj.columns))
            self.add_result(self.sheet,
                            self.row + month_report_obj.shape[0] + 2, *res)

            for i, cell in enumerate(self.sheet[self.row]):
                if i < 2:
                    continue
                cell.value = f'{cell.value:%d.%m}'

            self.style_report(row_start=self.row,
                              table_size=(month_report_obj.shape[0] + 3,
                                          month_report_obj.shape[1] + 1))

            margin = table_shift[['Date', 'month', 'Price']].groupby(
                ['Date', 'month']).agg('sum').reset_index()
            done = table_done[['Date', 'month', 'ResultOfShift_total']].groupby(
                ['Date', 'month']).agg('sum').reset_index()

            res_margin = [(1, 'ИТОГО ВЫРУЧКА'),
                          (2, sum(margin['Price']))]
            res_done = [(1, 'ВЫПЛАЧЕНО ЗАКАЗЧИКОМ'),
                        (2, sum(done['ResultOfShift_total']))]
            for i, date in enumerate(month_report_obj.columns[1:]):
                if date in list(margin['Date']):
                    res_margin += [(i + 3, margin[
                        margin['Date'] == date]['Price'].values[0])]
                else:
                    res_margin += [(i + 3, 0)]

                if date in list(done['Date']):
                    res_done += [(i + 3, done[
                        done['Date'] == date]['ResultOfShift_total'].values[0])]
                else:
                    res_done += [(i + 3, 0)]

            self.add_result(self.sheet, self.row + month_report_obj.shape[0] + 3, *res_margin)
            self.style_result(row=self.sheet[self.row + month_report_obj.shape[0] + 3],
                              column_count=month_report_obj.shape[1] + 1,
                              last_res=True)
            self.add_result(self.sheet, self.row + month_report_obj.shape[0] + 4, *res_done)
            self.style_result(row=self.sheet[self.row + month_report_obj.shape[0] + 4],
                              column_count=month_report_obj.shape[1] + 1,
                              last_res=True)

        self.work_book.save(self.file_name)

    def style_report(self, row_start, **kwargs):
        """
        Формат таблицы с табелем: сб, вс другим цветом, границы, формат числа и
        даты, стартовая и итоговая строки
        :param row_start: номер стартовой строки
        :param kwargs: table_size: (кол-во строк, кол-во столбцов). По умолчанию (1, 1).
        is_formatting: форматировать значения ячеек как числа и даты? По умолчанию - True.
        is_result_style: форматировать ли стартовую и итоговую строки таблицы? По умолчанию (True, True)
        :return: None
        """
        self.style_table(row_start=row_start, **kwargs)

        if 'table_size' not in kwargs:
            kwargs['table_size'] = (1, 1)
        row_end = row_start + kwargs['table_size'][0] - 1
        column_count = kwargs['table_size'][1]

        for i in range(2):
            self.sheet.merge_cells(start_row=row_start, start_column=i + 1,
                                   end_row=row_start + 1, end_column=i + 1)
            self.sheet.cell(row=row_start, column=i + 1).alignment = \
                Alignment(vertical='center', horizontal='center')

        for row in range(row_start, row_end + 1):
            for i, cell in enumerate(self.sheet[row]):
                if i >= column_count:
                    break
                if i == 1:
                    cell.fill = PatternFill(start_color='FED280',
                                            end_color='FED280',
                                            fill_type='solid')
                if self.sheet.cell(
                        row=row_start + 1, column=i + 1).value in ['сб', 'вс']:
                    if row < row_start + 2:
                        cell.font = Font(color='FF0000')
                    cell.fill = constants.REST_FILL

        self.style_auto_dimension(column_count=2)
        self.style_dimension(width=7,
                             column_numbers=list(range(3, column_count + 1)))


class WriteMarginTables(WriteTable):
    def write_gross_record(self, table, month=None):
        """
        Запись в Excel файл информационных данных и таблицы по городам по валовой прибыли
        :param table: ссылка на таблицу с валовой прибылью
        :param month: номер месяца
        :return: None
        """
        self.work_book = openpyxl.load_workbook(self.file_name)
        if month is not None:
            sheet_name = f'{self.first_sheet_name}_month_{month}'
            self.work_book.create_sheet(sheet_name)
        else:
            sheet_name = self.first_sheet_name

        self.sheet = self.work_book[sheet_name]
        self.row = 1

        with self.get_writer() as writer:
            writer.book = self.work_book
            writer.sheets = self.sheet_names
            self.write_titles(type_table=3, month=month)

            self.__write_margin_res(writer, table)

            self.row += 1
            city = 'Тюмень'
            self.__write_margin_res(writer, table, city=(city, 'ТМН'))

            city = 'Екатеринбург'
            self.__write_margin_res(writer, table, city=(city, 'ЕКБ'))

        self.work_book.save(self.file_name)

    def __write_margin_res(self, writer, table, city=None):
        """
        Запись в Excel файл итогов таблицы по городам по валовой прибыли
        или самой таблицы
        :param table: ссылка на таблицу с валовой прибылью
        :param writer: ссылка на writer
        :param city: (наименование города, краткое наименование города)
        :return: номер последней строки с результатом
        """
        if city is not None:
            city_name = city[0]
            short_city = city[1]

            self.row += 1
            res = ((1, f'ИТОГО {short_city}'),
                   (2, sum(table[table['Город'] == city_name]['Себестоимость'])),
                   (3, sum(table[table['Город'] == city_name]['Выручка'])),
                   (4, sum(table[table['Город'] == city_name]['Валовая прибыль'])),
                   (5, sum(table[table['Город'] == city_name]['Выплачено'])))
            self.add_result(self.sheet, self.row, *res)
            self.style_result(column_count=table.shape[1] - 1, last_res=True)

            self.row += 1
            diff = sum(table[table['Город'] == city_name]['Себестоимость']) - \
                   sum(table[table['Город'] == city_name]['Выплачено'])
            if sum(table[table['Город'] == city_name]['Выручка']) == 0:
                percentage = 0
            else:
                percentage = sum(table[table['Город'] == city_name]['Валовая прибыль']) / \
                             sum(table[table['Город'] == city_name]['Выручка'])
            res = ((1, f'ИТОГО - ВЫДАНО {short_city}'),
                   (2, diff),
                   (4, percentage))
            self.add_result(self.sheet, self.row, *res)
            self.style_result(column_count=table.shape[1] - 1, last_res=True)
            self.sheet.cell(row=self.row, column=4).value = \
                f'{self.sheet.cell(row=self.row, column=4).value * 100:.1f}%'
        else:
            columns_write = ['Заказчик', 'Себестоимость', 'Выручка',
                             'Валовая прибыль', 'Выплачено']
            res = ((1, 'ИТОГО'),
                   (2, sum(table['Себестоимость'])),
                   (3, sum(table['Выручка'])),
                   (4, sum(table['Валовая прибыль'])),
                   (5, sum(table['Выплачено'])))
            self.write_table(writer, table[columns_write], table_res=res,
                             is_auto_dimension=True)
