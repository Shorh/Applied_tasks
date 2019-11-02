import pandas as pd

from openpyxl.styles import PatternFill, Font, Border, Side


FONT_TITLE = Font(bold=True, size=14, color='0C4561')
RESULT_FILL = PatternFill(start_color='BDE9FF', end_color='BDE9FF',
                          fill_type='solid')
REST_FILL = PatternFill(start_color='DFDFDF', end_color='DFDFDF',
                        fill_type='solid')
NUMBER_FORMAT = '#,##0;\\-#,##0;;@'
DATA_FORMAT = 'mm-dd-yy'

today = pd.Timestamp.today()
TODAY = pd.Timestamp(year=today.year, month=today.month, day=today.day)
START_LAST_WEEK = TODAY + pd.Timedelta(days=-TODAY.weekday(), weeks=-2)
END_LAST_WEEK = START_LAST_WEEK + pd.Timedelta(days=13)
LAST_DAY_FOR_THE_FINE = START_LAST_WEEK + pd.Timedelta(days=18)

MONTH_START = START_LAST_WEEK.month
MONTH_END = END_LAST_WEEK.month
YEAR_START = START_LAST_WEEK.year
YEAR_END = END_LAST_WEEK.year

MONTHS = ['', 'ЯНВАРЬ', 'ФЕВРАЛЬ', 'МАРТ',
          'АПРЕЛЬ', 'МАЙ', 'ИЮНЬ',
          'ИЮЛЬ', 'АВГУСТ', 'СЕНТЯБРЬ',
          'ОКТЯБРЬ', 'НОЯБРЬ', 'ДЕКАБРЬ']

WEEKDAYS = ['вс', 'пн', 'вт', 'ср', 'чт', 'пт', 'сб']

THIN = Side(border_style="thin", color="000000")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
