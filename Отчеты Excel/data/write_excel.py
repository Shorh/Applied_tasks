import openpyxl
import pandas as pd
import numpy as np

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

FONT_TITLE = Font(bold=True, size=14, color='0C4561')
RESULT_FILL = PatternFill(start_color='BDE9FF', end_color='BDE9FF',
                          fill_type='solid')
REST_FILL = PatternFill(start_color='DFDFDF', end_color='DFDFDF',
                        fill_type='solid')
NUMBER_FORMAT = '#,##0;\\-#,##0;;@'
DATA_FORMAT = 'mm-dd-yy'

today = pd.Timestamp.today()
TODAY = pd.Timestamp(year=today.year, month=today.month, day=today.day)
MONTH = (TODAY - pd.Timedelta(days=3)).month
YEAR = (TODAY - pd.Timedelta(days=3)).year

MONTHS = ['', 'ЯНВАРЬ', 'ФЕВРАЛЬ', 'МАРТ',
          'АПРЕЛЬ', 'МАЙ', 'ИЮНЬ',
          'ИЮЛЬ', 'АВГУСТ', 'СЕНТЯБРЬ',
          'ОКТЯБРЬ', 'НОЯБРЬ', 'ДЕКАБРЬ',]

START_LAST_WEEK = TODAY + pd.Timedelta(days=-TODAY.weekday(), weeks=-1)
END_LAST_WEEK = START_LAST_WEEK + pd.Timedelta(days=6)

THIN = Side(border_style="thin", color="000000")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def auto_dimension(sheet, n_row=0, formatting=True):
    column_widths = []
    for i, row in enumerate(sheet):
        if i < n_row:
            continue
        for j, cell in enumerate(row):
            if formatting:
                cell.number_format = NUMBER_FORMAT
            if len(column_widths) > j:
                if cell.value and len(str(cell.value)) > column_widths[j]:
                    column_widths[j] = len(str(cell.value))
            else:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(i + 1)].width = column_width + 2


def add_result(sheet, res_row, *args):
    for column, value in args:
        sheet.cell(row=res_row, column=column).value = value


def result_style(row, column_count, last_res=False, alignment='center',
                 formatting=True):
    for i, cell in enumerate(row):
        if i >= column_count:
            break
        if formatting:
            cell.number_format = NUMBER_FORMAT
        cell.fill = RESULT_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
        if alignment:
            cell.alignment = Alignment(horizontal=alignment)
    if last_res:
        row[0].alignment = Alignment(horizontal='right')


def table_record(table_writer, table_sheet, table, table_res, table_name=None,
                 row=0, is_auto_dimension=False):
    if table.shape[0] != 0:
        if table_name:
            row += 2
            table_sheet.cell(row=row, column=1).value = table_name
            table_sheet.cell(row=row, column=1).font = FONT_TITLE
        row += 1

        table.to_excel(table_writer, table_sheet.title, index=False,
                       startrow=row - 1)
        if is_auto_dimension:
            auto_dimension(sheet=table_sheet, n_row=row - 1)

        row += table.shape[0] + 1
        add_result(table_sheet, row, *table_res)

        table_style(table_sheet=table_sheet, row_start=row - table.shape[0] - 1,
                    row_end=row, column_count=table.shape[1],
                    is_result_style=[True, True])
    return row


def create_month_report(table, obj):
    table_obj = pd.pivot_table(table[table['ObjectName'] == obj], index='FIO',
                               values='WorkedHours',
                               columns=['Date'], fill_value=0, aggfunc=np.sum)

    cur_day = 1
    day = pd.Timestamp(year=YEAR, month=MONTH, day=cur_day)
    while day.month == MONTH:
        if day not in list(table_obj.columns):
            table_obj[day] = pd.Series()
        day = day + pd.Timedelta(days=1)

    table_obj.fillna(0, inplace=True)
    table_obj.sort_index(axis=1, inplace=True)

    table_obj['ВСЕГО'] = table_obj.loc[:, :].apply(np.sum, axis=1)
    columns = ['ВСЕГО'] + list(table_obj.columns)[:-1]
    table_obj = table_obj[columns]

    return table_obj


def table_style(table_sheet, row_start, row_end, column_count,
                formatting=True, is_result_style=None):
    if is_result_style is None:
        is_result_style = [False, False]
    for row in range(row_start, row_end + 1):
        for i, cell in enumerate(table_sheet[row]):
            if i >= column_count:
                break
            if formatting:
                if 'Дата' in table_sheet.cell(row=row_start,
                                              column=i + 1).value:
                    cell.number_format = DATA_FORMAT
                else:
                    cell.number_format = NUMBER_FORMAT
            cell.border = BORDER
    if is_result_style[0]:
        result_style(table_sheet[row_start], column_count,
                     formatting=formatting)
    if is_result_style[1]:
        result_style(table_sheet[row_end], column_count, last_res=True,
                     formatting=formatting)


def table_report_style(table_sheet, row_start, row_end, column_count,
                       formatting=True, is_result_style=None):
    table_style(table_sheet=table_sheet, row_start=row_start, row_end=row_end,
                column_count=column_count, formatting=formatting,
                is_result_style=is_result_style)
    for i in range(2):
        table_sheet.merge_cells(start_row=row_start, start_column=i + 1,
                                end_row=row_start + 1, end_column=i + 1)
        table_sheet.cell(row=row_start, column=i + 1).alignment = \
            Alignment(vertical='center', horizontal='center')

    for row in range(row_start, row_end + 1):
        for i, cell in enumerate(table_sheet[row]):
            if i >= column_count:
                break
            if i == 1:
                cell.fill = PatternFill(start_color='FED280',
                                        end_color='FED280', fill_type='solid')
            if table_sheet.cell(row=row_start + 1, column=i + 1).value in \
                    ['сб', 'вс']:
                if row < row_start + 2:
                    cell.font = Font(color='FF0000')
                cell.fill = REST_FILL


def margin_res(sheet, row, table, writer,
               is_city=False, city=None, short_city=None):
    if is_city:
        row += 1
        res = ((1, f'ИТОГО {short_city}'),
               (2, sum(table[table['Город'] == city]['Себестоимость'])),
               (3, sum(table[table['Город'] == city]['Выручка'])),
               (4, sum(table[table['Город'] == city]['Валовая прибыль'])),
               (5, sum(table[table['Город'] == city]['Выплачено'])))
        add_result(sheet, row, *res)
        result_style(sheet[row], table.shape[1] - 1, last_res=True)

        row += 1
        diff = sum(table[table['Город'] == city]['Себестоимость']) - \
               sum(table[table['Город'] == city]['Выплачено'])
        if sum(table[table['Город'] == city]['Выручка']) == 0:
            percentage = 0
        else:
            percentage = sum(table[table['Город'] == city]['Валовая прибыль']) / \
                         sum(table[table['Город'] == city]['Выручка'])
        res = ((1, f'ИТОГО - ВЫДАНО {short_city}'),
               (2, diff),
               (4, percentage))
        add_result(sheet, row, *res)
        result_style(sheet[row], table.shape[1] - 1, last_res=True)
        sheet.cell(row=row,
                   column=4).value = f'{sheet.cell(row=row, column=4).value * 100:.1f}%'
    else:
        columns_write = ['Заказчик', 'Себестоимость', 'Выручка',
                         'Валовая прибыль', 'Выплачено']
        res = ((1, 'ИТОГО'),
               (2, sum(table['Себестоимость'])),
               (3, sum(table['Выручка'])),
               (4, sum(table['Валовая прибыль'])),
               (5, sum(table['Выплачено'])))
        row = table_record(writer, sheet, table[columns_write], res, row=row,
                           is_auto_dimension=True)

    return row


def gross_record(table, sheet, file_name, work_book, is_month=False):
    row = 1
    sheet.cell(row=row, column=1).value = f'ВАЛОВАЯ ПРИБЫЛЬ'
    if is_month:
        sheet.cell(row=row, column=1).value += f'. {MONTHS[MONTH]}'
    sheet.cell(row=row, column=1).font = FONT_TITLE
    row += 1
    sheet.cell(row=row, column=1).value = f'Дата формирования отчета: '
    sheet.cell(row=row, column=2).value = f'{today:%d.%m.%Y}'
    row += 1
    if not is_month:
        sheet.cell(row=row, column=1).value = f'Отчетные даты: '
        sheet.cell(row=row,
                   column=2).value = f'{START_LAST_WEEK:%d.%m.%Y} - ' \
                                     f'{END_LAST_WEEK:%d.%m.%Y}'
        row += 1

    for i, _ in enumerate(sheet):
        sheet.cell(row=i + 1, column=1).alignment = Alignment(horizontal='left')

    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a',
                        datetime_format='DD/MM') as writer:
        writer.book = work_book
        writer.sheets = dict((ws.title, ws) for ws in work_book.worksheets)

        row = margin_res(sheet, row, table, writer)

        row += 1
        city = 'Тюмень'
        row = margin_res(sheet, row, table, writer, is_city=True, city=city,
                         short_city='ТМН')

        city = 'Екатеринбург'
        row = margin_res(sheet, row, table, writer, is_city=True, city=city,
                         short_city='ЕКБ')

    return row
