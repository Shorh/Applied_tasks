import openpyxl

from openpyxl.styles import PatternFill, Font, Alignment
from data import constants


def auto_dimension(sheet, row_count=0, add_width=2, is_formatting=True):
    """
    Автоширина столбцов таблицы в файле Excel
    :param sheet: ссылка на страницу Excel
    :param row_count: количество строк
    :param add_width: добавить к ширине символов. По умолчанию - 2
    :param is_formatting: форматировать значения ячеек как числа?
    :return: None
    """
    column_widths = []
    for i, row in enumerate(sheet):
        if i < row_count:
            continue
        for j, cell in enumerate(row):
            if is_formatting:
                cell.number_format = constants.NUMBER_FORMAT
            if len(column_widths) > j:
                if cell.value and len(str(cell.value)) > column_widths[j]:
                    column_widths[j] = len(str(cell.value))
            else:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(i + 1)].width = column_width + add_width


def result_style(row, column_count, last_res=False,
                 alignment='center', is_formatting=True):
    """
    Форматирование стартовых и итоговых строк таблицы
    :param row: ссылка на строку Excel
    :param column_count: количество столбцов
    :param last_res: является ли строка итоговой? По умолчанию - False
    :param alignment: выравнивание, по умолчанию - центр
    :param is_formatting: форматировать значения ячеек как числа? По умолчанию - True
    :return: None
    """
    for i, cell in enumerate(row):
        if i >= column_count:
            break
        if is_formatting:
            cell.number_format = constants.NUMBER_FORMAT
        cell.fill = constants.RESULT_FILL
        cell.font = Font(bold=True)
        cell.border = constants.BORDER
        if alignment:
            cell.alignment = Alignment(horizontal=alignment)
    if last_res:
        row[0].alignment = Alignment(horizontal='right')


def table_style(sheet, row_start, table_size=(1, 1),
                is_formatting=True, is_result_style=None):
    """
    Формат таблицы: границы, формат числа и даты, стартовая и итоговая строки
    :param sheet: ссылка на страницу Excel
    :param row_start: номер стартовой строки
    :param table_size: (кол-во строк, кол-во столбцов). По умолчанию - (1, 1)
    :param is_formatting: форматировать значения ячеек как числа и даты? По умолчанию - False
    :param is_result_style: форматировать ли стартовую и итоговую строки таблицы? По умолчанию - (False, False)
    :return: None
    """
    row_end = row_start + table_size[0] - 1
    column_count = table_size[1]

    if is_result_style is None:
        is_result_style = (False, False)
    for row in range(row_start, row_end + 1):
        for i, cell in enumerate(sheet[row]):
            if i >= column_count:
                break
            if is_formatting:
                if 'Дата' in sheet.cell(row=row_start, column=i + 1).value:
                    cell.number_format = constants.DATA_FORMAT
                else:
                    cell.number_format = constants.NUMBER_FORMAT
            cell.border = constants.BORDER
    if is_result_style[0]:
        result_style(sheet[row_start], column_count,
                     is_formatting=is_formatting)
    if is_result_style[1]:
        result_style(sheet[row_end], column_count,
                     last_res=True, is_formatting=is_formatting)


def table_report_style(sheet, row_start, table_size=(1, 1),
                       is_formatting=True, is_result_style=None):
    """
    Формат таблицы с табелем: сб, вс другим цветом, границы, формат числа и
    даты, стартовая и итоговая строки
    :param sheet: ссылка на страницу Excel
    :param row_start: номер стартовой строки
    :param table_size: (кол-во строк, кол-во столбцов). По умолчанию - (1, 1)
    :param is_formatting: форматировать значения ячеек как числа и даты? По умолчанию - False
    :param is_result_style: форматировать ли стартовую и итоговую строки таблицы? По умолчанию - (False, False)
    :return: None
    """
    table_style(sheet=sheet, row_start=row_start, table_size=table_size,
                is_formatting=is_formatting, is_result_style=is_result_style)

    row_end = row_start + table_size[0] - 1
    column_count = table_size[1]

    for i in range(2):
        sheet.merge_cells(start_row=row_start, start_column=i + 1,
                          end_row=row_start + 1, end_column=i + 1)
        sheet.cell(row=row_start, column=i + 1).alignment = \
            Alignment(vertical='center', horizontal='center')

    for row in range(row_start, row_end + 1):
        for i, cell in enumerate(sheet[row]):
            if i >= column_count:
                break
            if i == 1:
                cell.fill = PatternFill(start_color='FED280',
                                        end_color='FED280', fill_type='solid')
            if sheet.cell(row=row_start + 1, column=i + 1).value in \
                    ['сб', 'вс']:
                if row < row_start + 2:
                    cell.font = Font(color='FF0000')
                cell.fill = constants.REST_FILL
